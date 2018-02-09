"""created on 2018.1.30
    author:deyakong
"""
import random
import datetime
import openpyxl


class ExcelCell:
    def __init__(self, value, cell_loc, is_sub_title=True, other_loc=None):
        self.value = value
        self.location = cell_loc
        if other_loc is not None:
            self.is_combine = True
            self.combine_loc = other_loc
        else:
            self.is_combine = False
            self.combine_loc = None
        self.is_sub_title = is_sub_title
        self.sub_cell = None

    def get_value(self):
        return self.value

    def get_loc(self):
        return self.location

    def get_combine_loc(self):
        return self.combine_loc

    def change_loc(self, new_loc):
        self.location = new_loc

    def add_sub_content(self, value):
        if self.is_sub_title:
            loc = self.location
            new_loc = loc[0] + str(int(loc[1])+1)
            self.sub_cell = ExcelCell(value, cell_loc=new_loc, is_sub_title=False)
        else:
            return


def generate_numbers():
    integer = random.randint(0, 100)
    decimal = integer+random.randint(0, 100)/100
    text_integer = str(integer)
    text_decimal = str(decimal)
    long_text_integer = '1000000000000000000000000000000000000000000000000000000000000000000000000000000000'
    long_text_decimal = '100000000000000000000000000000.1111111111111111111111111111111111111111111111111111'
    negative = -random.randint(0, 100)
    negative_decimal = -(integer+random.randint(0, 100)/100)
    negative_long_text_integer = '-1000000000000000000000000000000000000000000000000000000000000000000000000000000000'
    negative_long_text_decimal = '-100000000000000000000000000000.1111111111111111111111111111111111111111111111111111'
    none_value = None
    letter_front = 'a123'
    letter_middle = '1a23'
    letter_after = '123a'
    space_front = ' 123'
    space_middle = '12 3'
    space_after = '123 '
    only_space = '   '
    zero = 0
    return_button = '12\n3'
    numbers_dic = {
        'integer': integer,
        'decimal': decimal,
        'text_integer': text_integer,
        'text_decimal': text_decimal,
        'long_text_integer': long_text_integer,
        'long_text_decimal': long_text_decimal,
        'negative': negative,
        'negative_decimal': negative_decimal,
        'negative_long_text_integer': negative_long_text_integer,
        'negative_long_text_decimal': negative_long_text_decimal,
        'none_value': none_value,
        'letter_front': letter_front,
        'letter_middle': letter_middle,
        'letter_after': letter_after,
        'space_front': space_front,
        'space_middle': space_middle,
        'space_after': space_after,
        'only_space': only_space,
        'zero': zero,
        'return_button': return_button
    }
    return numbers_dic


def generate_chinese_text(limit=(0, 100)):
    a_long_str = """山海经》是知识的山，是知识的海，并以它广博、丰富的内容和奇特、高超的想象力为古往今来的人们所称道、所
    叹服，因而它不仅是广大社会科学和自然科学工作者研究的重要对象，而且也是广大读者朋友获得许多古代文化、历史、民俗等知识
    的宝库。此书所记的山、水、国、民族、动物、植物、矿物、药物等，除大部分是殊异的而外，也有一部分是常见的。然而，无论其
    是殊异的还是常见的，都是上古历史、地理、风俗的一个侧面。所以，要了解古代的山川地理、民俗风物、奇兽怪鸟、神仙魔鬼、金
    玉珍宝、自然矿物、神话故事，不可不读 《山海经》。而且，《山海经》也为人们了解自然知识和古代某些充满巫神祈祷的社会生
    活提供了宝贵的资料。尤其是《山海经》中所保存的为人们所熟知的精卫填海、夸父追日、羿射九日、禹鲧治水、共工怒触不周山等
    神话传说，不仅是以幻想的形式反映了人与自然的矛盾，更重要的是给人们以积极鼓舞；而那些如太阳每天的东升西落、月亮每月的
    圆缺盈亏、一年四季的寒暑变化等有关自然界的记述，不仅仅是上古人要为解释自然现象而作，更重要的是给人们以探索的勇气。诸
    如此类，真可谓有永久的魅力！然而，由于《山海经》既非出自一人之手，也非写成于一时，文字记载上本已疏略简乱，再加上流传
    时间久远，所以，讹误衍脱、增削窜改的情况很多，使人不便阅读，往往难以理解，需要做一些校勘疏通的工作。\n
    但考虑到本书为一部普及性读物，在于简明，不宜进行繁琐考证，故采取通便之法，既能保持原文面貌，又可校正理顺文字，还要不
    必列出校勘记。这就是：凡遇讹文，即标以圆括号，后面补上改正之字，并标以方括号，表示更正，如《南山经》鹊山条中的“其状
    如穀而黑理”一句，“穀”为“榖”之讹，于是校正为“其状如（穀）[榖]而黑理”；凡遇衍文，即标以圆括号，表示删除，如《南
    山经》最后一条中的“一璧稻米”一句，在上下之中读不通，实为衍文，于是校正为“（一璧稻米）”；凡遇脱文，即标以方括号，
    表示增补，如 《南山经》柢山条中的“又东三百里柢山”一句，“三百里”下脱一“曰”字，于是校正为“又东三百里[曰]柢山”。
    凡作如此校正的文字，都是有可靠根据的，是在吸收古今学者研究成果的基础上进行的，如明人王崇庆，清人王念孙、何焯、吴任臣
    、汪绂、毕沅、郝懿行，今人汪绍原、袁珂等，其中尤多得益于袁珂的校勘成果。又《山海经》中每多异读字、通假字，难识难读，
    故随文用拼音注上音，而且不避重复，以方便今天的青年读者。本书的注释本着通俗易懂的原则，避免繁征博引，力图简明扼要，以
    疏通文义为主。注释中所据成说，主要出自晋人郭璞的《山海经传》、清人郝懿行的《山海经笺疏》、今人袁珂的《山海经校注》。
    另外，也间出己意，以发明前贤所通，却难决断，则于注释中列出，以供读者鉴别择取。如遇意义不明，又无旁证可求之处，则仍存
    其疑，不做臆解，以求慎重。除个别情况外，不做重复注释。本书的译文，以直译为主，以意译为辅，尽量避免以释代译。 未逮者
    。如遇似是而非之说，或几种说法均可"""
    uper_limit = a_long_str[:limit[1]]
    over_uper_limit = a_long_str[:limit[1]+1]
    lower_limit = a_long_str[:limit[0]]
    over_lower_limit = a_long_str[:limit[0]-1]
    conformed = a_long_str[:limit[1]-limit[0]+1]
    chinese_text_dic = {
        'uper_limit': uper_limit,
        'lower_limit': lower_limit,
        'over_uper_limit': over_uper_limit,
        'over_lower_limit': over_lower_limit,
        'conformed': conformed
    }
    return chinese_text_dic


def generate_date(year=None, month=None, day=None):
    date_date = datetime.datetime(year, month, day)
    text_date_minus = date_date.strftime('%Y-%m-%d')
    text_date_backslash = date_date.strftime('%Y/%m/%d')
    unexpected_month = '2018-13-01'
    unexpected_day = '2018-13-32'
    unexpected_leep_month = '2018-2-29'
    date_dic = {
        'date_date': date_date,
        'text_date_minus': text_date_minus,
        'text_date_backslash': text_date_backslash,
        'unexpected_month': unexpected_month,
        'unexpected_leep_month': unexpected_leep_month,
        'unexpected_day': unexpected_day
    }
    return date_dic


def create_teachplan_excel():
    def generate_excel(title, course_list, other_info_list):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells('A1:D1')
        ws['A1'] = '班主任教学计划表'
        ws['A2'] = '培训班名称'
        ws['A3'] = '培训对象'
        ws['A4'] = '培训人数'
        ws['A5'] = '办班开始时间'
        ws['A6'] = '办班结束时间'
        ws['A7'] = '培训学时'
        ws['A8'] = '培训地点'
        ws['A9'] = '承办单位'
        ws['A10'] = '课程名称'
        ws['B10'] = '学时'
        ws['C10'] = '拟聘请主讲人员姓名'
        ws['D10'] = '主讲人单位'
        ws['A' + str(11+len(course_list))] = '交流、研讨主题'
        ws['C' + str(11 + len(course_list))] = '学时'
        ws['A' + str(12 + len(course_list))] = '现场教学地点'
        ws['C' + str(12 + len(course_list))] = '学时'
        ws['A' + str(13 + len(course_list))] = '其它学时'
        ws['A' + str(14 + len(course_list))] = '联系人'
        ws['C' + str(14 + len(course_list))] = '联系电话'
        for i in range(0, 7):
            row = str(i+3)
            ws.merge_cells('B'+row+':D'+row)
            ws['B' + row] = other_info_list[i]
        for j in range(len(course_list)):
            ws['A' + str(11+j)] = course_list[j][0]
            ws['B' + str(11 + j)] = course_list[j][1]
            ws['C' + str(11 + j)] = course_list[j][2]
            ws['D' + str(11 + j)] = course_list[j][3]
            ws['B' + str(11 + len(course_list))] = other_info_list[7]
            ws['D' + str(11 + len(course_list))] = other_info_list[8]
            ws['B' + str(12 + len(course_list))] = other_info_list[9]
            ws['D' + str(12 + len(course_list))] = other_info_list[10]
            ws.merge_cells('B'+str(13 + len(course_list))+':D'+str(13 + len(course_list)))
            ws['B' + str(13 + len(course_list))] = other_info_list[11]
            ws['B' + str(14 + len(course_list))] = other_info_list[12]
            ws['D' + str(14 + len(course_list))] = other_info_list[13]
        wb.save('D:\\测试项目\\林业局二期\\生成\\'+title+'.xlsx')
    target_list = generate_chinese_text((1, 100))
    amount_list = generate_numbers()
    start_time_list = generate_date(2018, 1, 30)
    end_time_list = generate_date(2018, 2, 1)
    total_hours_list = generate_numbers()
    place_list = generate_chinese_text((1, 300))
    company_list = generate_chinese_text((1, 100))
    course_name_list = generate_chinese_text((1, 50))
    course_hours_list = generate_numbers()
    teacher_name_list = generate_chinese_text((1, 50))
    teacher_company_list = generate_chinese_text((1, 50))
    topic_list = generate_chinese_text((1, 300))
    communicate_hours_list = generate_numbers()
    live_place_list = generate_chinese_text((1, 300))
    live_hours_list = generate_numbers()
    other_hours = generate_numbers()
    contact_list = generate_chinese_text((1, 100))
    phone = '13222112321'

    """generate_excel('teaching_plan_temple', [[course_name_list['conformed'], course_hours_list['integer'],
                                             teacher_name_list['conformed'], teacher_company_list['conformed']]],
                   [target_list['conformed'], amount_list['integer'], start_time_list['date_date'],
                    end_time_list['date_date'], total_hours_list['integer'], place_list['conformed'],
                    company_list['conformed'], topic_list['conformed'], communicate_hours_list['integer'],
                    live_place_list['conformed'], live_hours_list['integer'], other_hours['integer'],
                    contact_list['conformed'], phone])"""

    for item in target_list:
        generate_excel('teaching_plan_target_'+item, [[course_name_list['conformed'], course_hours_list['integer'],
                                                       teacher_name_list['conformed'], teacher_company_list['conformed']
                                                       ]],
                       [target_list[item], amount_list['integer'], start_time_list['date_date'],
                        end_time_list['date_date'], total_hours_list['integer'], place_list['conformed'],
                        company_list['conformed'], topic_list['conformed'], communicate_hours_list['integer'],
                        live_place_list['conformed'], live_hours_list['integer'], other_hours['integer'],
                        contact_list['conformed'], phone])
    for item in amount_list:
        generate_excel('teaching_plan_amount_'+item, [[course_name_list['conformed'], course_hours_list['integer'],
                                                       teacher_name_list['conformed'], teacher_company_list['conformed']
                                                       ]],
                       [target_list['conformed'], amount_list[item], start_time_list['date_date'],
                        end_time_list['date_date'], total_hours_list['integer'], place_list['conformed'],
                        company_list['conformed'], topic_list['conformed'], communicate_hours_list['integer'],
                        live_place_list['conformed'], live_hours_list['integer'], other_hours['integer'],
                        contact_list['conformed'], phone])
    for item in start_time_list:
        generate_excel('teaching_plan_starttime_'+item, [[course_name_list['conformed'], course_hours_list['integer'],
                                                          teacher_name_list['conformed'],
                                                          teacher_company_list['conformed']]],
                       [target_list['conformed'], amount_list['integer'], start_time_list[item],
                        end_time_list['date_date'], total_hours_list['integer'], place_list['conformed'],
                        company_list['conformed'], topic_list['conformed'], communicate_hours_list['integer'],
                        live_place_list['conformed'], live_hours_list['integer'], other_hours['integer'],
                        contact_list['conformed'], phone])
    for item in end_time_list:
        generate_excel('teaching_plan_endtime_'+item, [[course_name_list['conformed'], course_hours_list['integer'],
                                                        teacher_name_list['conformed'],
                                                        teacher_company_list['conformed']]],
                       [target_list['conformed'], amount_list['integer'], start_time_list['date_date'],
                        end_time_list[item], total_hours_list['integer'], place_list['conformed'],
                        company_list['conformed'], topic_list['conformed'], communicate_hours_list['integer'],
                        live_place_list['conformed'], live_hours_list['integer'], other_hours['integer'],
                        contact_list['conformed'], phone])
    for item in total_hours_list:
        generate_excel('teaching_plan_total_hours_'+item, [[course_name_list['conformed'], course_hours_list['integer'],
                                                            teacher_name_list['conformed'],
                                                            teacher_company_list['conformed']]],
                       [target_list['conformed'], amount_list['integer'], start_time_list['date_date'],
                        end_time_list['date_date'], total_hours_list[item], place_list['conformed'],
                        company_list['conformed'], topic_list['conformed'], communicate_hours_list['integer'],
                        live_place_list['conformed'], live_hours_list['integer'], other_hours['integer'],
                        contact_list['conformed'], phone])
    for item in place_list:
        generate_excel('teaching_plan_place_'+item, [[course_name_list['conformed'], course_hours_list['integer'],
                                                      teacher_name_list['conformed'], teacher_company_list['conformed']
                                                      ]],
                       [target_list['conformed'], amount_list['integer'], start_time_list['date_date'],
                        end_time_list['date_date'], total_hours_list['integer'], place_list[item],
                        company_list['conformed'], topic_list['conformed'], communicate_hours_list['integer'],
                        live_place_list['conformed'], live_hours_list['integer'], other_hours['integer'],
                        contact_list['conformed'], phone])
    for item in company_list:
        generate_excel('teaching_plan_company_'+item, [[course_name_list['conformed'], course_hours_list['integer'],
                                                        teacher_name_list['conformed'],
                                                        teacher_company_list['conformed']]],
                       [target_list['conformed'], amount_list['integer'], start_time_list['date_date'],
                        end_time_list['date_date'], total_hours_list['integer'], place_list['conformed'],
                        company_list[item], topic_list['conformed'], communicate_hours_list['integer'],
                        live_place_list['conformed'], live_hours_list['integer'], other_hours['integer'],
                        contact_list['conformed'], phone])
    for item in topic_list:
        generate_excel('teaching_plan_topic_'+item, [[course_name_list['conformed'], course_hours_list['integer'],
                                                      teacher_name_list['conformed'], teacher_company_list['conformed']
                                                      ]],
                       [target_list['conformed'], amount_list['integer'], start_time_list['date_date'],
                        end_time_list['date_date'], total_hours_list['integer'], place_list['conformed'],
                        company_list['conformed'], topic_list[item], communicate_hours_list['integer'],
                        live_place_list['conformed'], live_hours_list['integer'], other_hours['integer'],
                        contact_list['conformed'], phone])
    for item in communicate_hours_list:
        generate_excel('teaching_plan_communicate_hours_'+item, [[course_name_list['conformed'],
                                                                  course_hours_list['integer'],
                                                                  teacher_name_list['conformed'],
                                                                  teacher_company_list['conformed']]],
                       [target_list['conformed'], amount_list['integer'], start_time_list['date_date'],
                        end_time_list['date_date'], total_hours_list['integer'], place_list['conformed'],
                        company_list['conformed'], topic_list['conformed'], communicate_hours_list[item],
                        live_place_list['conformed'], live_hours_list['integer'], other_hours['integer'],
                        contact_list['conformed'], phone])
    for item in live_place_list:
        generate_excel('teaching_plan_live_place_'+item, [[course_name_list['conformed'], course_hours_list['integer'],
                                                           teacher_name_list['conformed'],
                                                           teacher_company_list['conformed']]],
                       [target_list['conformed'], amount_list['integer'], start_time_list['date_date'],
                        end_time_list['date_date'], total_hours_list['integer'], place_list['conformed'],
                        company_list['conformed'], topic_list['conformed'], communicate_hours_list['integer'],
                        live_place_list[item], live_hours_list['integer'], other_hours['integer'],
                        contact_list['conformed'], phone])
    for item in live_hours_list:
        generate_excel('teaching_plan_live_hours_'+item, [[course_name_list['conformed'], course_hours_list['integer'],
                                                           teacher_name_list['conformed'],
                                                           teacher_company_list['conformed']]],
                       [target_list['conformed'], amount_list['integer'], start_time_list['date_date'],
                        end_time_list['date_date'], total_hours_list['integer'], place_list['conformed'],
                        company_list['conformed'], topic_list['conformed'], communicate_hours_list['integer'],
                        live_place_list['conformed'], live_hours_list[item], other_hours['integer'],
                        contact_list['conformed'], phone])
    for item in other_hours:
        generate_excel('teaching_plan_other_hours_'+item, [[course_name_list['conformed'], course_hours_list['integer'],
                                                            teacher_name_list['conformed'],
                                                            teacher_company_list['conformed']]],
                       [target_list['conformed'], amount_list['integer'], start_time_list['date_date'],
                        end_time_list['date_date'], total_hours_list['integer'], place_list['conformed'],
                        company_list['conformed'], topic_list['conformed'], communicate_hours_list['integer'],
                        live_place_list['conformed'], live_hours_list['integer'], other_hours[item],
                        contact_list['conformed'], phone])
    for item in contact_list:
        generate_excel('teaching_plan_contact_'+item, [[course_name_list['conformed'], course_hours_list['integer'],
                                                        teacher_name_list['conformed'],
                                                        teacher_company_list['conformed']]],
                       [target_list['conformed'], amount_list['integer'], start_time_list['date_date'],
                        end_time_list['date_date'], total_hours_list['integer'], place_list['conformed'],
                        company_list['conformed'], topic_list['conformed'], communicate_hours_list['integer'],
                        live_place_list['conformed'], live_hours_list['integer'], other_hours['integer'],
                        contact_list[item], phone])
    for item in course_name_list:
        generate_excel('teaching_plan_course_name_'+item, [[course_name_list[item], course_hours_list['integer'],
                                                            teacher_name_list['conformed'],
                                                            teacher_company_list['conformed']]],
                       [target_list['conformed'], amount_list['integer'], start_time_list['date_date'],
                        end_time_list['date_date'], total_hours_list['integer'], place_list['conformed'],
                        company_list['conformed'], topic_list['conformed'], communicate_hours_list['integer'],
                        live_place_list['conformed'], live_hours_list['integer'], other_hours['integer'],
                        contact_list['conformed'], phone])
    for item in course_hours_list:
        generate_excel('teaching_plan_course_hours_'+item, [[course_name_list['conformed'], course_hours_list[item],
                                                            teacher_name_list['conformed'],
                                                            teacher_company_list['conformed']]],
                       [target_list['conformed'], amount_list['integer'], start_time_list['date_date'],
                        end_time_list['date_date'], total_hours_list['integer'], place_list['conformed'],
                        company_list['conformed'], topic_list['conformed'], communicate_hours_list['integer'],
                        live_place_list['conformed'], live_hours_list['integer'], other_hours['integer'],
                        contact_list['conformed'], phone])
    for item in teacher_name_list:
        generate_excel('teaching_plan_teacher_name_'+item, [[course_name_list['conformed'],
                                                             course_hours_list['integer'],
                                                             teacher_name_list[item],
                                                             teacher_company_list['conformed']]],
                       [target_list['conformed'], amount_list['integer'], start_time_list['date_date'],
                        end_time_list['date_date'], total_hours_list['integer'], place_list['conformed'],
                        company_list['conformed'], topic_list['conformed'], communicate_hours_list['integer'],
                        live_place_list['conformed'], live_hours_list['integer'], other_hours['integer'],
                        contact_list['conformed'], phone])
    for item in teacher_company_list:
        generate_excel('teaching_plan_teacher_company_'+item, [[course_name_list['conformed'],
                                                                course_hours_list['integer'],
                                                                teacher_name_list['conformed'],
                                                                teacher_company_list[item]]],
                       [target_list['conformed'], amount_list['integer'], start_time_list['date_date'],
                        end_time_list['date_date'], total_hours_list['integer'], place_list['conformed'],
                        company_list['conformed'], topic_list['conformed'], communicate_hours_list['integer'],
                        live_place_list['conformed'], live_hours_list['integer'], other_hours['integer'],
                        contact_list['conformed'], phone])
create_teachplan_excel()



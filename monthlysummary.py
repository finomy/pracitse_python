'''
Created on 2017年5月3日

@author: dykong
'''
import os
import os.path
import re
import time

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.reader.excel import load_workbook


class project:
    def __init__(self,name):
        self.missions=[]
        self.project=name
        self.time=1
        self.thinmissions=[]
        self.jobs=[]
        self.thinjobs=[]
    def addmission(self,mission1):
        self.missions.append(mission1)
    def getproject(self):
        return self.project

    def getmissions(self):
#        for i in range(len(self.missions)):
#            print(self.missions[i])
        return self.missions
    def getmission(self,index):
        return self.missions[index]
    def getmissionlength(self):
        return len(self.missions)
    def deletemission(self,index):
        self.missions.pop(index)
    def timeplus(self,num):
        self.time=self.time+num
    def gettime(self):
        return self.time
    def addthinmissions(self,missions):
        self.thinmissions=missions
    def getthinmission(self,index):
        return self.thinmissions[index]
    def getthinmissionlength(self):
        return len(self.thinmissions)
    def getjob(self,i):
        return self.jobs[i]
    def getjobs(self):
        return self.jobs
    def addjob(self,job):
        self.jobs.append(job)
    def addjobs(self,jobs):
        self.jobs=self.jobs+jobs
    def setjob(self,jobs):
        self.jobs=jobs
    def addthinjobs(self,jobs):
        self.thinjobs=jobs
    def getthinjob(self,i):
        return self.thinjobs[i]
    def getthinjobs(self):
        return self.thinjobs
def writetxt():
    file=open('triteplans.txt','w')
    file.writelines('总天数：'+str(len(fitfiles))+'\n')
    for i in range(len(projects)):
        totaljobs=''
        file.writelines(projects[i].getproject()+':('+str(projects[i].gettime())+')'+'\n')
        file.writelines("---------------------------------------------"+"\n")  
        file.writelines('任务：\n')
        for j in range(projects[i].getthinmissionlength()):
            file.writelines(projects[i].getthinmission(j)+'\n')
        file.writelines("---------------------------------------------"+"\n") 
        file.writelines('每天完成的工作汇总：\n')    
        for k in range(len(projects[i].getthinjobs())):
            totaljobs=totaljobs+projects[i].getthinjob(k)+'。'
        totaljobs=totaljobs+'\n'
        file.writelines(totaljobs)
        file.writelines("*********************************************"+"\n")   
    file.close()
def writelog(content):
    file=open('log.txt','a')
    file.writelines
    localtime=time.strftime("%Y %m %d %H:%M:%S",time.localtime())
    file.writelines(localtime+' '+content+'\r\n')
    file.close()
def get_property():
    file=open('msconfig.txt','r')
    lines=file.readlines()
    return lines
    

lines=get_property()
rootdir=re.sub('\n','',lines[0])
usr=re.sub('\n','',lines[1])
month=re.sub('\n','',lines[2])
filedir=os.listdir(rootdir)
fitfiles=[]
projects=[]
rootdir2=rootdir+'\\'
for allDir in filedir:
    child = os.path.join('%s%s' % (rootdir2,allDir))
#    print(child)
    if re.match('.*'+usr+'2018年'+month+'月.*日报.xlsx',child)!=None:
        fitfiles.append(child)
        print(child)
print(len(fitfiles))
for i in range(len(fitfiles)):
    wb=load_workbook(fitfiles[i])
    print(fitfiles[i])
    writelog('deal with '+fitfiles[i])
    ws=wb.active
    B2=ws['B2']
    B3=ws['B3']
#    print(B2)
    B2value=B2.value
    B3value=B3.value
#    print(B2value)
    lines=B2value.split('\n')
    lines.pop(0)
    lines2=B3value.split('\n')
    lines2.pop(0)
    one_day_projects=[]
    today_job=[]
    NoneProjectFlag=0
    for j in range(len(lines)):   #get plans in one excel
#        print(i,' ',j,' ',lines[j])
        
        if re.match('.*项目：',lines[j])!=None:
            words=lines[j]
#            print(words)
            words=re.sub('\d\.','',words)
            words=re.sub('\w\.','',words)
            words=re.sub('\w\、','',words)
            words=re.sub('\d\、','',words)
            words=re.sub('\：','',words)
            words=re.sub('\:','',words)
            words=re.sub(' ','',words)
            words=re.sub('\r','',words)
            words=re.sub('\n','',words)
            print(words)
            odp=project(words)
            one_day_projects.append(odp)
#            print('enter one day report project')
        elif re.match('.*项目:',lines[j])!=None:
            words=lines[j]
#            print(words)
            words=re.sub('\d\.','',words)
            words=re.sub('\w\.','',words)
            words=re.sub('\w\、','',words)
            words=re.sub('\d\、','',words)
            words=re.sub('\：','',words)
            words=re.sub('\:','',words)
            words=re.sub(' ','',words)
            words=re.sub('\r','',words)
            words=re.sub('\n','',words)
            print(words)
            odp=project(words)
            one_day_projects.append(odp)
            
#            print('enter one day report project')
                   
        elif lines[j]!=None:
            print(lines[j])
            a=lines[j]
            a=re.sub('\d\.','',a)
            a=re.sub('\(\d*%\)','',a)
            a=re.sub('\(.*\)','',a)
            a=re.sub('\（.*\）','',a)
            a=re.sub(' ','',a)
            if len(a)==0:
                continue
            one_day_projects[len(one_day_projects)-1].addmission(a)
        
#    print('FLAG:',NoneProjectFlag)
    if(len(one_day_projects)==0):
        NoneProjectFlag=1
    
#    print('FLAG:',NoneProjectFlag)
    if NoneProjectFlag==1:    
        one_day_projects.append(project('计划外'))
#            print(lines[j][2:])
#            print('enter one day report mission')
    for p in range(len(lines2)):
        jobcontent=re.findall('\d.(\w*)',lines2[p])
#        print(len(jobcontent))
        if(len(jobcontent)==0):
#            print(lines2[p])
            continue
            
        else:
            today_job.append(jobcontent[0])
    if len(projects)==0:
            for o in range(len(one_day_projects)):
#                print('add to projects:'+one_day_projects[o].getproject())
                projects.append(one_day_projects[o]) #if projects list is empty,add above plans in projects list
#                print('add to projects:'+projects[0].getproject())
            projects[0].setjob(today_job)    
    else:
        for j in range(len(one_day_projects)):
#            print('add to projects:'+one_day_projects[j].getproject())
            flag_is_exist=-1
            for k in range(len(projects)):
                if projects[k].getproject()==one_day_projects[j].getproject():
                    for l in range(one_day_projects[j].getmissionlength()):
                        
                        projects[k].addmission(one_day_projects[j].getmission(l))
                    projects[k].timeplus(1)
                    projects[k].addjobs(today_job)
                    flag_is_exist=0
                    
                    break
            if flag_is_exist==-1:
                projects.append(one_day_projects[j])
                projects[len(projects)-1].setjob(today_job)
#                print('add to projects:'+projects[len(projects)-1].getproject())
            

for i in range(len(projects)):
#    projects[i].getmissions()
    dic={}
    print(projects[i].getproject())
    for j in range(projects[i].getmissionlength()):        
        if dic.get(projects[i].getmission(j))==None:
            dic[projects[i].getmission(j)]=1
        else:
            dic[projects[i].getmission(j)]=dic.get(projects[i].getmission(j))+1
#    print(dic)
    projects[i].addthinmissions(missions=tuple(dic.keys()))
    print(dic)
for i in range(len(projects)):
    div={}
    for j in range(len(projects[i].getjobs())):
        if dic.get(projects[i].getjob(j))==None:
            dic[projects[i].getjob(j)]=1
        else:
            dic[projects[i].getjob(j)]=dic.get(projects[i].getjob(j))+1
#    print(dic)
    projects[i].addthinjobs(tuple(dic.keys()))
    print('完成工作：')
    print(projects[i].getproject())
    print(dic)
writetxt()